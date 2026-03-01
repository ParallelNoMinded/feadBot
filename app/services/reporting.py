import asyncio
import io
from datetime import datetime
from typing import Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.utils.hotel_timezone import convert_to_timezone
from shared_models import (
    Comment,
    Feedback,
    FeedbackComment,
    Hotel,
    User,
    UserHotel,
    Zone,
)


class ReportingService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def export_xlsx(
        self,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        hotels_scope: Optional[Sequence[str]] = None,
    ) -> bytes:
        # Build base query joining related tables including comments
        stmt = (
            select(
                Feedback,
                User,
                Zone,
                Comment,
                Hotel.timezone,
                UserHotel.room_number,
                UserHotel.first_name,
                UserHotel.last_name,
                UserHotel.open,
                UserHotel.close,
            )
            .join(UserHotel, UserHotel.id == Feedback.user_stay_id)
            .join(User, User.id == UserHotel.user_id)
            .join(Zone, Zone.id == Feedback.zone_id)
            .join(Hotel, Hotel.id == Zone.hotel_id)
            .outerjoin(FeedbackComment, FeedbackComment.feedback_id == Feedback.id)
            .outerjoin(Comment, Comment.id == FeedbackComment.comment_id)
        )
        # Enforce hotel scope if provided (including empty list -> no access)
        if hotels_scope is not None:
            if len(hotels_scope) == 0:
                # Return empty XLSX with headers only
                wb = Workbook()
                ws = wb.active
                headers = [
                    "Telegram ID",
                    "Номер телефона",
                    "Номер апартамента",
                    "Дата отзыва",
                    "Зона",
                    "Оценка",
                    "Комментарий",
                    "Фамилия",
                    "Имя",
                    "Дата заезда",
                    "Дата выезда",
                ]
                ws.append(headers)
                # Make headers bold
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()
            stmt = stmt.where(Hotel.short_name.in_(list(hotels_scope)))
        if date_from is not None:
            try:
                stmt = stmt.where(Feedback.created_at >= date_from)
            except Exception:
                pass
        if date_to is not None:
            try:
                stmt = stmt.where(Feedback.created_at <= date_to)
            except Exception:
                pass
        rows = (await self.session.execute(stmt)).all()

        # Run in thread to avoid blocking
        xlsx_bytes = await asyncio.to_thread(self._build_xlsx_from_rows, rows)
        return xlsx_bytes

    @staticmethod
    def _build_xlsx_from_rows(rows) -> bytes:
        """Build XLSX file from database rows (CPU-bound, runs in thread pool)"""
        wb = Workbook()
        ws = wb.active
        # Russian headers per requirements
        headers = [
            "Telegram ID",
            "Номер телефона",
            "Номер апартамента",
            "Дата отзыва",
            "Зона",
            "Оценка",
            "Комментарий",
            "Фамилия",
            "Имя",
            "Дата заезда",
            "Дата выезда",
        ]
        ws.append(headers)
        # Make headers bold
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Group rows by feedback to handle multiple comments per feedback
        feedback_data = {}
        for fb, user, zone, comment, hotel_timezone, room_number, first_name, last_name, open, close in rows:
            feedback_id = fb.id
            if feedback_id not in feedback_data:
                feedback_data[feedback_id] = {
                    "feedback": fb,
                    "user": user,
                    "zone": zone,
                    "hotel_timezone": hotel_timezone,
                    "room_number": room_number or "",
                    "comments": [],
                    "first_name": first_name or "",
                    "last_name": last_name or "",
                    "open": open.strftime("%d.%m.%Y") if open else "",
                    "close": close.strftime("%d.%m.%Y") if close else "",
                }
            if comment:
                feedback_data[feedback_id]["comments"].append(comment.comment)

        # Write data rows
        for data in feedback_data.values():
            # Join all comments with newline separator
            comments_text = "\n".join(data["comments"]) if data["comments"] else ""
            created_at = convert_to_timezone(data["feedback"].created_at, data["hotel_timezone"])
            ws.append(
                [
                    data["user"].external_user_id,
                    data["user"].phone_number,
                    data["room_number"],
                    created_at.strftime("%d.%m.%Y %H:%M"),
                    data["zone"].name,
                    data["feedback"].rating or "",
                    comments_text,
                    data["first_name"],
                    data["last_name"],
                    data["open"],
                    data["close"],
                ]
            )

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
